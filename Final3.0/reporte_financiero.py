import openpyxl


class ReporteFinanciero:
    def generar_informe(self, excel_filename):
        # Crear un libro de trabajo de Excel
        wb = openpyxl.Workbook()
        ws = wb.active

        # Agregar encabezados a las columnas
        ws.append(["Fecha", "Descripción", "Categoría", "Monto"])

        # Agregar las transacciones a las filas del archivo Excel
        for transaccion in self.transacciones:
            ws.append([transaccion.fecha, transaccion.descripcion, transaccion.categoria, transaccion.monto])

        # Guardar el archivo Excel
        wb.save(excel_filename)

    def guardar_transacciones_excel(self, excel_filename, transacciones):
        try:
            wb = openpyxl.load_workbook(excel_filename)
        except FileNotFoundError:
            # Si el archivo no existe, crea uno nuevo
            wb = openpyxl.Workbook()

        ws = wb.active

        # Agrega las transacciones a la hoja de trabajo
        for transaccion in transacciones:
            ws.append([transaccion.fecha, transaccion.descripcion, transaccion.categoria, transaccion.monto,
                       transaccion.tipo])

        # Guarda el archivo
        wb.save(excel_filename)

    def generar_informe_txt(self, txt_filename, excel_transacciones_filename, excel_metas_ahorro_filename):
        # Leer datos de transacciones desde el archivo Excel
        transacciones = []
        wb_transacciones = openpyxl.load_workbook(excel_transacciones_filename)
        ws_transacciones = wb_transacciones.active
        for row in ws_transacciones.iter_rows(min_row=2, values_only=True):
            fecha, descripcion, categoria, monto, tipo = row
            transacciones.append((fecha, descripcion, categoria, monto, tipo))

        # Leer datos de metas de ahorro desde el archivo Excel
        metas_ahorro = []
        wb_metas_ahorro = openpyxl.load_workbook(excel_metas_ahorro_filename)
        ws_metas_ahorro = wb_metas_ahorro.active
        for row in ws_metas_ahorro.iter_rows(min_row=2, values_only=True):
            descripcion, monto_objetivo, fecha_limite = row
            metas_ahorro.append((descripcion, monto_objetivo, fecha_limite))

        with open(txt_filename, 'w') as txt_file:
            txt_file.write("Información de Transacciones:\n")
            for transaccion in transacciones:
                txt_file.write(
                    f"Fecha: {transaccion[0]}, Descripción: {transaccion[1]}, Categoría: {transaccion[2]}, Monto: {transaccion[3]}, Tipo: {transaccion[4]}\n")

            txt_file.write("\nInformación de Metas de Ahorro:\n")
            for meta in metas_ahorro:
                txt_file.write(f"Descripción: {meta[0]}, Monto Objetivo: {meta[1]}, Fecha Límite: {meta[2]}\n")
