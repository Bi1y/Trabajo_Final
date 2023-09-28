import openpyxl
from transaccion import Transaccion
from meta_ahorro import MetaAhorro
import os


class Usuario:
    def __init__(self, nombre):
        self.nombre = nombre
        self.transacciones = []
        self.metas_ahorro = []

    def establecer_meta_ahorro(self, descripcion, monto_objetivo, fecha_limite):
        meta = MetaAhorro(descripcion, monto_objetivo, fecha_limite)
        self.metas_ahorro.append(meta)
        self.guardar_metas_ahorro_excel("data/metas_ahorro.xlsx")

    def guardar_metas_ahorro_excel(self, excel_filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Descripción", "Monto Objetivo", "Fecha Límite"])

        for meta in self.metas_ahorro:
            ws.append([meta.descripcion, meta.monto_objetivo, meta.fecha_limite])

        wb.save(excel_filename)

    def agregar_transaccion(self, fecha, descripcion, categoria, monto, tipo):
        excel_filename = "data/transacciones.xlsx"

        if not os.path.exists(excel_filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Fecha", "Descripción", "Categoría", "Monto", "Tipo"])
            wb.save(excel_filename)

        transaccion = Transaccion(fecha, descripcion, categoria, monto, tipo)
        self.transacciones.append(transaccion)

        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active
        ws.append([fecha, descripcion, categoria, monto, tipo])
        wb.save(excel_filename)

    def eliminar_transaccion(self, transaccion):
        self.transacciones.remove(transaccion)

    def obtener_resumen_financiero(self):
        total_gastos = 0
        total_ingresos = 0
        excel_filename = "data/transacciones.xlsx"

        try:
            wb_transacciones = openpyxl.load_workbook(excel_filename)
            ws_transacciones = wb_transacciones.active
        except FileNotFoundError:
            return {
                "total_gastos": 0,
                "total_ingresos": 0,
                "balance": 0,
            }

        for row in ws_transacciones.iter_rows(min_row=2, values_only=True):
            fecha, descripcion, categoria, monto, tipo = row
            if tipo == "gasto":
                total_gastos += monto
            elif tipo == "ingreso":
                total_ingresos += monto

        balance = total_ingresos - total_gastos

        return {
            "total_gastos": total_gastos,
            "total_ingresos": total_ingresos,
            "balance": balance,
        }

    def ver_historial_transacciones(self):
        excel_transacciones_filename = "data/transacciones.xlsx"

        try:
            wb_transacciones = openpyxl.load_workbook(excel_transacciones_filename)
            ws_transacciones = wb_transacciones.active

            if ws_transacciones.max_row < 2:
                print("El archivo de transacciones no contiene datos.")
                return

            print("\nHistorial de Transacciones:")
            for row in ws_transacciones.iter_rows(min_row=2, values_only=True):
                fecha, descripcion, categoria, monto, tipo = row
                print(
                    f"Fecha: {fecha}, Descripción: {descripcion}, Categoría: {categoria}, Monto: ${monto:.2f}, Tipo: {tipo}")
        except FileNotFoundError:
            print("No hay transacciones")
        except Exception as e:
            print(f"Se produjo un error al leer el archivo de transacciones: {str(e)}")

    def generar_informe_financiero(self, txt_filename):
        excel_transacciones_filename = "data/transacciones.xlsx"

        if os.path.exists(excel_transacciones_filename):
            transacciones = []
            wb_transacciones = openpyxl.load_workbook(excel_transacciones_filename)
            ws_transacciones = wb_transacciones.active
            for row in ws_transacciones.iter_rows(min_row=2, values_only=True):
                fecha, descripcion, categoria, monto, tipo = row
                transacciones.append((fecha, descripcion, categoria, monto, tipo))
        else:
            transacciones = None

        excel_metas_ahorro_filename = "data/metas_ahorro.xlsx"

        if os.path.exists(excel_metas_ahorro_filename):
            metas_ahorro = []
            wb_metas_ahorro = openpyxl.load_workbook(excel_metas_ahorro_filename)
            ws_metas_ahorro = wb_metas_ahorro.active
            for row in ws_metas_ahorro.iter_rows(min_row=2, values_only=True):
                descripcion, monto_objetivo, fecha_limite = row
                metas_ahorro.append((descripcion, monto_objetivo, fecha_limite))
        else:
            metas_ahorro = None

        with open(txt_filename, 'a') as txt_file:
            if transacciones is not None:
                txt_file.write("Información de Transacciones:\n")
                for transaccion in transacciones:
                    txt_file.write(
                        f"Fecha: {transaccion[0]}, Descripción: {transaccion[1]}, Categoría: {transaccion[2]}, Monto: {transaccion[3]}, Tipo: {transaccion[4]}\n")
                txt_file.write("\n")

            if metas_ahorro is not None:
                txt_file.write("Información de Metas de Ahorro:\n")
                for meta in metas_ahorro:
                    txt_file.write(f"Descripción: {meta[0]}, Monto Objetivo: {meta[1]}, Fecha Límite: {meta[2]}\n")
